#!/usr/bin/python3
# -*- coding: utf-8

import pymysql
import openpyxl
import datetime
from dateutil.relativedelta import relativedelta
import calendar
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill
from openpyxl.utils.cell import get_column_letter
import os
import yaml
from yaml import CLoader as Loader, CDumper as Dumper

TABLE_HEAD = ['Договор (дата)', 'Заказ (дата)', 'Абонентская (руб)', 'За превыш. (руб)', 'Модем (тариф)',
              'Передано (Мб)', 'Принято (Мб)', 'Перерасход', 'Итого']

RUSSIAN_MONTH = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь',
                 7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
# Directory with files
DIRECTORY = os.path.dirname(os.path.abspath(__file__))

#Config file
with open(f'{DIRECTORY}/config.yaml'.format(DIRECTORY=DIRECTORY), 'r') as config:
    config_data = yaml.load(config, Loader=Loader)

# Billing
HOST_BILLING = config_data['billing']['host']
PORT_BILLING = config_data['billing']['port']
USER_NAME_BILLING = config_data['billing']['user']
PASSWORD_BILLING = config_data['billing']['password']
DATABASE_BILLING = config_data['billing']['db']

# RADIUS
HOST_RADIUS = config_data['radius']['host']
PORT_RADIUS = config_data['radius']['port']
USER_NAME_RADIUS = config_data['radius']['user']
PASSWORD_RADIUS = config_data['radius']['password']
DATABASE_RADIUS = config_data['radius']['db']

def make_custom_column_width(sheet_name, start_column, end_column, custom_width):
    col_xl = [get_column_letter(idx) for idx in range(1, 16384+1)] #16384 is Excel's limit
    for col in col_xl[col_xl.index(start_column):col_xl.index(end_column)+1]:
        sheet_name.column_dimensions[col].width = custom_width

def make_custom_style(work_sheet, work_style, span):
    for cell_tuple in work_sheet[span]:
        for cell in cell_tuple:
            cell.style = work_style

##### Генерация Даты ######
current_date = datetime.datetime.now() 
previous_month_date = current_date + relativedelta(months = -1) #Вычитаем из текущей даты месяц
PREVIOUS_YEAR_MONTH = previous_month_date.strftime('%Y-%m')
LAST_DAY_PREVIOUS_MONTH = calendar.monthrange(previous_month_date.year, previous_month_date.month)[1]
FIRST_DAY_QUERY = f'{PREVIOUS_YEAR_MONTH}-01 00:00:00'
LAST_DAY_QUERY = f'{PREVIOUS_YEAR_MONTH}-{LAST_DAY_PREVIOUS_MONTH} 23:59:59'

clients_info = {}

try:
##### Соединение с локальной БД #####
    with pymysql.connect(
        host=HOST_BILLING, port=PORT_BILLING, user=USER_NAME_BILLING, password=PASSWORD_BILLING,
        database=DATABASE_BILLING, cursorclass=pymysql.cursors.DictCursor
        ) as connection_local_db:
        with connection_local_db.cursor() as cursor_local_db:
            cursor_local_db.execute('SELECT client_id, company, username\
                                     FROM clients')
            for client in cursor_local_db.fetchall():
                client_id = int(client.get('client_id'))
                clients_info[client_id] = {}
                clients_info[client_id]['company_name'] = client.get('company')
                clients_info[client_id]['username'] = client.get('username')
except OSError as ex:
    print('Проблемы соединения с локальной БД!')
    print(ex)


##### Соединение с БД Radius #####
try:
    with pymysql.connect(
        host=HOST_RADIUS, port=PORT_RADIUS, user=USER_NAME_RADIUS, password=PASSWORD_RADIUS,
        database=DATABASE_RADIUS, cursorclass=pymysql.cursors.DictCursor
        ) as connection_radius_db:
            with connection_radius_db.cursor() as cursor_radius_db:
                for client_id, client_data in clients_info.items():
                    ##### Переданные данные #####
                    cursor_radius_db.execute('SELECT SUM(acctinputoctets) \
                                              FROM radacct \
                                              WHERE calledstationid=(%s)\
                                              AND acctstarttime\
                                              BETWEEN (%s) AND (%s)',\
                                              (client_data['username'],FIRST_DAY_QUERY, LAST_DAY_QUERY))
                    data_transmit_bytes = cursor_radius_db.fetchone()
                    transmit_bytes = data_transmit_bytes.get('SUM(acctinputoctets)')
                    transmit_bytes = int(transmit_bytes) if transmit_bytes != None else 0
                    clients_info[client_id]['output'] = round(transmit_bytes / 1024 / 1024, 3)

                    ##### Загруженные данные #####
                    cursor_radius_db.execute('SELECT SUM(acctoutputoctets) \
                                              FROM radacct \
                                              WHERE calledstationid=(%s) AND acctstarttime >= (%s)', \
                                              (client_data['username'], previous_month_date))
                    data_recieve_bytes = cursor_radius_db.fetchone()
                    recieve_bytes = data_recieve_bytes.get('SUM(acctoutputoctets)')
                    recieve_bytes = int(recieve_bytes) if recieve_bytes != None else 0
                    clients_info[client_id]['input'] = round(recieve_bytes / 1024 / 1024, 3)

except Exception as ex:
    print('Проблемы соединения с базой данных RADIUS!')
    print(ex)


##### Creating XLS file #####

###### Initialization first list of book #####
wb = openpyxl.Workbook()
##### Creating of title style #####
title_style = NamedStyle(name='title')
title_style.font = Font(bold=True, size=11)
title_border = Side(style='thin', color='000000')
title_style.border = Border(left=title_border, top=title_border, right=title_border, bottom=title_border)
title_style.alignment = Alignment(horizontal='center', vertical='center')
title_style.fill = PatternFill('solid', fgColor='d9d9d9')

##### Creating usual style #####
usual_style = NamedStyle(name='usual')
usual_style.font = Font(size=11)
usual_border = Side(style='thin', color='000000')
usual_style.border = Border(left=usual_border, top=usual_border, right=usual_border, bottom=usual_border)
usual_style.alignment = Alignment(horizontal='center', vertical='center')

##### Initialization of companies lists #####
for client_id, client_data in clients_info.items() :
    company_sheet = wb.create_sheet(str(client_data['company_name']))
    make_custom_column_width(company_sheet, 'B', 'J', 20)
    make_custom_style(company_sheet, title_style, 'B2:D2')
    company_sheet['B2'] = RUSSIAN_MONTH[previous_month_date.month]
    company_sheet['C2'] = previous_month_date.year
    company_sheet['D2'] = client_data['company_name']

##### Creating head of table #####
    for iteration_number, name in enumerate(TABLE_HEAD):
        company_sheet.cell(row=5, column=iteration_number+2).value = name
        company_sheet.cell(row=5, column=iteration_number+2).style = title_style


##### Connecting to local DB and filling of table #####
    try:
        connection_billing = pymysql.connect(
        host=HOST_BILLING, port=PORT_BILLING, user=USER_NAME_BILLING, password=PASSWORD_BILLING,
        database=DATABASE_BILLING, cursorclass=pymysql.cursors.DictCursor)
        cursor_billing = connection_billing.cursor()
        #print('Successfully connected!')

        ##### Number and date of contract #####
        contracts_info = {}
        cursor_billing.execute('SELECT contract_id, number, date\
                                FROM contracts\
                                WHERE client_id=(%s)',\
                                (client_id))
        data_contract_number = cursor_billing.fetchall()
        for contract in data_contract_number:
            contract_id = int(contract.get('contract_id'))
            contracts_info[contract_id] = {}
            contracts_info[contract_id]['number'] = contract.get('number')
            contracts_info[contract_id]['date'] = contract.get('date').strftime('%d.%m.%Y')

        for iteration_number, contract_id in enumerate(contracts_info.keys()):
            row_number = iteration_number + 6
            make_custom_style(company_sheet, usual_style, f'B{row_number}:J{row_number}')
            company_sheet[f'B{row_number}'] = f"{contracts_info[contract_id]['number']} ({contracts_info[contract_id]['date']})"

            ##### Number and date of order #####
            cursor_billing.execute('SELECT number, date, cost, modem_sn, c_over\
                                    FROM orders\
                                    WHERE orders.contract_id=(%s)', contract_id)
            data_orders = cursor_billing.fetchall()
            order_number = data_orders[0].get('number')
            date_order = data_orders[0].get('date').strftime('%d.%m.%Y')
            order_cost = data_orders[0].get('cost')
            order_cost_over = data_orders[0].get('c_over')
            modem_sn = data_orders[0].get('modem_sn')
            ##### Tarif limit of clients #####
            cursor_billing.execute('SELECT tarif\
                                    FROM limit_client\
                                    WHERE modem_sn=(%s)', modem_sn)
            data_tarif = cursor_billing.fetchall()
            limit = data_tarif[0].get('tarif') if isinstance(data_tarif, list) else 0
            ##### Filling of table #####
            company_sheet[f'C{row_number}'] = f"{order_number} ({date_order})"
            company_sheet[f'D{row_number}'] = order_cost
            company_sheet[f'E{row_number}'] = order_cost_over if order_cost_over != 0 else '-'
            company_sheet[f'F{row_number}'] = f'{modem_sn} ({limit})' if modem_sn != 0 else '-'
            company_sheet[f'G{row_number}'] = client_data['output'] if modem_sn != 0 else '-'
            company_sheet[f'H{row_number}'] = client_data['input'] if modem_sn != 0 else '-'
            
            ##### Calculation of money #####
            sum_trafic = client_data['input'] + client_data['output']
            if limit != 0 and sum_trafic > limit:
                    company_sheet[f'I{row_number}'] = sum_trafic - limit
                    company_sheet[f'J{row_number}'] = order_cost + (sum_trafic - limit) * order_cost_over
            else:
                company_sheet[f'I{row_number}'] = '-'
                company_sheet[f'J{row_number}'] = order_cost

    except OSError as ex:
        print('Проблемы соединения с локальной БД!')
        print(ex)
    finally:
        connection_billing.close()
        cursor_billing.close()

del wb['Sheet']

PATH = '/usr/traffic-counter/month_report/reports/'
filename = f'{previous_month_date.strftime("%B-%Y")}.xlsx'
wb.save(PATH + filename)
print('Well done!')

